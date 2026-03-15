"""
utils/data_reader.py
─────────────────────
Reads test data from test_data.xlsx.
  - DataReader.config(key)    → reads from ⚙️ Config sheet
  - DataReader.tc(tc_id, key) → reads from TC-specific sheet
  - DataReader.all_tc(tc_id)  → returns full dict for a TC sheet
"""

from __future__ import annotations
from pathlib import Path
from functools import lru_cache
import sys

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from config.config import DATA_DIR

EXCEL_PATH = DATA_DIR / "test_data.xlsx"
CONFIG_SHEET = "⚙️ Config"

# Data row start (rows 1-2 = banner, 3 = section hdr, 4-8 = cfg rows, 9+ = more)
_DATA_START_ROW = 4   # first key/value row in Config sheet
_TC_DATA_START  = 11  # first data row in TC sheets (after headers at row 10)


class DataReader:
    """Singleton-style cached Excel reader."""

    _wb = None

    @classmethod
    def _get_wb(cls):
        if cls._wb is None:
            if not EXCEL_PATH.exists():
                raise FileNotFoundError(
                    f"test_data.xlsx not found at {EXCEL_PATH}\n"
                    f"Run: python gen_test_data_excel.py  to generate it."
                )
            cls._wb = load_workbook(str(EXCEL_PATH), data_only=True)
        return cls._wb

    # ── Config sheet ──────────────────────────────────────────────────────────
    @classmethod
    def config(cls, key: str, default: str = "") -> str:
        """Read a value from the ⚙️ Config sheet by key name."""
        wb  = cls._get_wb()
        ws  = wb[CONFIG_SHEET]
        for row in ws.iter_rows(min_row=_DATA_START_ROW, values_only=True):
            if row[0] and str(row[0]).strip() == key:
                return str(row[1]).strip() if row[1] is not None else default
        return default

    @classmethod
    def all_config(cls) -> dict[str, str]:
        """Return all key-value pairs from the Config sheet."""
        wb  = cls._get_wb()
        ws  = wb[CONFIG_SHEET]
        return {
            str(r[0]).strip(): str(r[1]).strip()
            for r in ws.iter_rows(min_row=_DATA_START_ROW, values_only=True)
            if r[0] and r[1] is not None
        }

    # ── TC sheets ─────────────────────────────────────────────────────────────
    @classmethod
    def tc(cls, tc_id: str, key: str, default: str = "") -> str:
        """Read a single value from a TC sheet by key name."""
        wb = cls._get_wb()
        if tc_id not in wb.sheetnames:
            raise ValueError(f"Sheet '{tc_id}' not found in test_data.xlsx")
        ws = wb[tc_id]
        for row in ws.iter_rows(min_row=_TC_DATA_START, values_only=True):
            if row[0] and str(row[0]).strip() == key:
                return str(row[1]).strip() if row[1] is not None else default
        return default

    @classmethod
    def all_tc(cls, tc_id: str) -> dict[str, str]:
        """Return all key-value pairs from a TC sheet as a dict."""
        wb = cls._get_wb()
        if tc_id not in wb.sheetnames:
            return {}
        ws = wb[tc_id]
        return {
            str(r[0]).strip(): str(r[1]).strip()
            for r in ws.iter_rows(min_row=_TC_DATA_START, values_only=True)
            if r[0] and r[1] is not None
        }

    @classmethod
    def reload(cls):
        """Force reload from disk (call if Excel was edited while running)."""
        cls._wb = None


# ── Convenience shortcuts ─────────────────────────────────────────────────────
def cfg(key: str, default: str = "") -> str:
    return DataReader.config(key, default)

def td(tc_id: str, key: str, default: str = "") -> str:
    return DataReader.tc(tc_id, key, default)
