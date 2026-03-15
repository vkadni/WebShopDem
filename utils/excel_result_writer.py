"""
utils/excel_result_writer.py
─────────────────────────────
Creates a NEW Excel result file after each run.
  - One sheet per TC
  - Each step row has: Step No, Step Name, Status, Timestamp, Notes, Screenshot (embedded image)
  - Summary sheet with overall results
  - JIRA import sheet is NEVER touched
"""

from __future__ import annotations
from datetime import datetime
from pathlib import Path
from typing import Optional
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from config.config import REPORT_DIR

REPORT_DIR.mkdir(parents=True, exist_ok=True)

thin  = Side(style="thin",   color="CCCCCC")
thick = Side(style="medium", color="1E4D8C")
THIN_B  = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
THICK_B = Border(left=thick, right=thick, top=thick, bottom=thick)

STATUS_STYLE = {
    "pass":        ("D4EDDA", "155724"),
    "Pass":        ("D4EDDA", "155724"),
    "fail":        ("F8D7DA", "721C24"),
    "Fail":        ("F8D7DA", "721C24"),
    "Not Run":     ("FFF3CD", "856404"),
    "Blocked":     ("E2E3E5", "383D41"),
    "In Progress": ("CCE5FF", "004085"),
}

def _fill(h): return PatternFill("solid", fgColor=h)
def _font(bold=False, color="000000", size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)
def _aln(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

IMG_W_PX  = 320   # screenshot width embedded in Excel (pixels)
IMG_H_PX  = 180   # screenshot height embedded in Excel (pixels)
ROW_H_SS  = 140   # row height in Excel units for screenshot rows


def _make_summary_sheet(wb: Workbook, results: list[dict]) -> None:
    ws = wb.active
    ws.title = " Summary"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "  DemoWebShop – Test Execution Result"
    ws["A1"].font  = _font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill  = _fill("1E4D8C")
    ws["A1"].alignment = _aln("left")
    ws.row_dimensions[1].height = 32

    run_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.merge_cells("A2:H2")
    ws["A2"] = f"  Generated: {run_time}  |  Tester: {results[0]['tester'] if results else ''}"
    ws["A2"].font  = _font(color="FFFFFF", size=9)
    ws["A2"].fill  = _fill("2E6DA4")
    ws["A2"].alignment = _aln("left")
    ws.row_dimensions[2].height = 16

    # Stats
    total   = len(results)
    passed  = sum(1 for r in results if r["status"] == "Pass")
    failed  = sum(1 for r in results if r["status"] == "Fail")
    not_run = sum(1 for r in results if r["status"] == "Not Run")
    blocked = sum(1 for r in results if r["status"] == "Blocked")

    ws.row_dimensions[3].height = 8
    stat_cols = [
        ("Total",   total,   "1E4D8C", "DDEEFF"),
        ("Pass",    passed,  "155724", "D4EDDA"),
        ("Fail",    failed,  "721C24", "F8D7DA"),
        ("Not Run", not_run, "856404", "FFF3CD"),
        ("Blocked", blocked, "383D41", "E2E3E5"),
    ]
    for ci, (label, val, fg, bg) in enumerate(stat_cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14
        lc = ws.cell(row=4, column=ci, value=label)
        lc.font = _font(bold=True, color="FFFFFF", size=10)
        lc.fill = _fill("1E4D8C"); lc.alignment = _aln("center"); lc.border = THIN_B
        vc = ws.cell(row=5, column=ci, value=val)
        vc.font = _font(bold=True, color=fg, size=20)
        vc.fill = _fill(bg); vc.alignment = _aln("center"); vc.border = THIN_B
        ws.row_dimensions[5].height = 36

    ws.row_dimensions[6].height = 10

    # Header row
    headers = ["TC ID","Summary","Module","Priority","Status",
               "Exec Date","Exec Time","Actual Result"]
    widths  = [12, 36, 26, 10, 12, 14, 12, 40]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=7, column=ci, value=h)
        c.font = _font(bold=True, color="FFFFFF", size=10)
        c.fill = _fill("1E4D8C"); c.alignment = _aln("center"); c.border = THIN_B
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[7].height = 20

    for ri, r in enumerate(results, 8):
        bg_hex, fg_hex = STATUS_STYLE.get(r["status"], ("FFFFFF","000000"))
        row_bg = "F8F9FA" if ri % 2 == 0 else "FFFFFF"
        vals = [r["tc_id"], r["summary"], r["module"], r["priority"],
                r["status"], r["exec_date"], r["exec_time"], r["actual_result"]]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = THIN_B
            c.alignment = _aln("center" if ci in (1,4,5,6,7) else "left",
                                wrap=ci == 8)
            if ci == 5:
                c.font = _font(bold=True, color=fg_hex, size=10)
                c.fill = _fill(bg_hex)
            elif ci == 1:
                c.font = _font(bold=True, color="1E4D8C", size=10)
                c.fill = _fill("D6E4F7")
            else:
                c.font = _font(size=9)
                c.fill = _fill(row_bg)
        ws.row_dimensions[ri].height = 22


def _make_tc_sheet(wb: Workbook, result: dict) -> None:
    tc_id   = result["tc_id"]
    ws      = wb.create_sheet(title=tc_id)
    ws.sheet_view.showGridLines = False

    # ── TC header ─────────────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    ws["A1"] = f"  {tc_id} – {result['summary']}"
    ws["A1"].font = _font(bold=True, color="FFFFFF", size=13)
    ws["A1"].fill = _fill("1E4D8C"); ws["A1"].alignment = _aln("left")
    ws.row_dimensions[1].height = 28

    # Metadata
    meta = [
        ("Module",         result["module"],       "Exec Date",  result["exec_date"]),
        ("Priority",       result["priority"],      "Exec Time",  result["exec_time"]),
        ("Status",         result["status"],        "Tester",     result["tester"]),
        ("Actual Result",  result["actual_result"], "Defect ID",  result.get("defect_id","")),
    ]
    for ri, (l1, v1, l2, v2) in enumerate(meta, 2):
        ws.row_dimensions[ri].height = 20
        for ci, (lbl, val) in enumerate([(l1,v1),(l2,v2)], 1):
            col_l = (ci - 1) * 3 + 1
            col_v = col_l + 1
            lc = ws.cell(row=ri, column=col_l, value=lbl)
            lc.font = _font(bold=True, color="FFFFFF", size=9)
            lc.fill = _fill("1E4D8C"); lc.alignment = _aln("right"); lc.border = THIN_B

            vc = ws.cell(row=ri, column=col_v, value=val)
            if lbl == "Status":
                bg_h, fg_h = STATUS_STYLE.get(val, ("FFFFFF","000000"))
                vc.font = _font(bold=True, color=fg_h, size=9)
                vc.fill = _fill(bg_h)
            else:
                vc.font = _font(size=9)
                vc.fill = _fill("EAF2FB")
            vc.alignment = _aln("left", wrap=True); vc.border = THIN_B
            ws.merge_cells(start_row=ri, start_column=col_v,
                           end_row=ri,   end_column=col_v+1)

    ws.row_dimensions[6].height = 8

    # ── Step table header ──────────────────────────────────────────────────────
    step_hdr_row = 7
    step_headers = ["Step","Step Description","Status","Time","Notes","Screenshot"]
    step_widths  = [6, 38, 12, 10, 28, 46]
    for ci, (h, w) in enumerate(zip(step_headers, step_widths), 1):
        c = ws.cell(row=step_hdr_row, column=ci, value=h)
        c.font = _font(bold=True, color="FFFFFF", size=10)
        c.fill = _fill("2E6DA4"); c.alignment = _aln("center"); c.border = THIN_B
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[step_hdr_row].height = 20

    # ── Step rows ──────────────────────────────────────────────────────────────
    for step in result.get("steps", []):
        row_idx = step_hdr_row + step["number"]
        ws.row_dimensions[row_idx].height = ROW_H_SS

        bg_hex, fg_hex = STATUS_STYLE.get(step["status"], ("FFFFFF","000000"))
        row_bg = "F8F9FA" if step["number"] % 2 == 0 else "FFFFFF"

        # Col A – Step number
        c = ws.cell(row=row_idx, column=1, value=step["number"])
        c.font = _font(bold=True, color="1E4D8C", size=11)
        c.fill = _fill("D6E4F7"); c.alignment = _aln("center"); c.border = THIN_B

        # Col B – Step name
        c = ws.cell(row=row_idx, column=2, value=step["name"])
        c.font = _font(size=9); c.fill = _fill(row_bg)
        c.alignment = _aln("left", wrap=True); c.border = THIN_B

        # Col C – Status
        c = ws.cell(row=row_idx, column=3, value=step["status"].upper())
        c.font = _font(bold=True, color=fg_hex, size=9)
        c.fill = _fill(bg_hex); c.alignment = _aln("center"); c.border = THIN_B

        # Col D – Timestamp
        c = ws.cell(row=row_idx, column=4, value=step["timestamp"])
        c.font = _font(size=9); c.fill = _fill(row_bg)
        c.alignment = _aln("center"); c.border = THIN_B

        # Col E – Notes
        c = ws.cell(row=row_idx, column=5, value=step.get("notes",""))
        c.font = _font(size=9); c.fill = _fill(row_bg)
        c.alignment = _aln("left", wrap=True); c.border = THIN_B

        # Col F – Screenshot (embedded image)
        ss_path = Path(step.get("screenshot_path",""))
        if ss_path.exists():
            try:
                img = XLImage(str(ss_path))
                img.width  = IMG_W_PX
                img.height = IMG_H_PX
                cell_addr  = f"F{row_idx}"
                ws.add_image(img, cell_addr)
            except Exception as ex:
                ws.cell(row=row_idx, column=6,
                        value=f"[Image error: {ex}]").font = _font(size=8, color="888888")
        else:
            ws.cell(row=row_idx, column=6,
                    value="[No screenshot]").font = _font(size=8, color="888888")

        # Add border to screenshot cell
        ws.cell(row=row_idx, column=6).border = THIN_B
        ws.cell(row=row_idx, column=6).fill   = _fill(row_bg)


def write_excel_results(results: list[dict]) -> Path:
    """
    Create a fresh Excel result file with:
      - Sheet 1: Summary
      - Sheet per TC: step-by-step table with embedded screenshots
    """
    wb = Workbook()
    _make_summary_sheet(wb, results)

    for r in results:
        _make_tc_sheet(wb, r)

    run_id   = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = REPORT_DIR / f"TestResult_{run_id}.xlsx"
    wb.save(str(out_path))
    print(f"    Excel result → {out_path}")
    return out_path
