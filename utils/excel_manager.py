"""
utils/excel_manager.py
──────────────────────
Reads test data FROM the Excel file and writes execution results BACK
into the JIRA-import sheet (columns J-P).
"""

from __future__ import annotations
from pathlib import Path
from datetime import datetime
from typing import Optional
import sys

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from config.config import EXCEL_RESULT_FILE

# ── Style helpers ─────────────────────────────────────────────────────────────
thin = Side(style="thin", color="CCCCCC")
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

STATUS_STYLE: dict[str, tuple[str, str]] = {
    "Pass":        ("D4EDDA", "155724"),
    "Fail":        ("F8D7DA", "721C24"),
    "Blocked":     ("E2E3E5", "383D41"),
    "Not Run":     ("FFF3CD", "856404"),
    "In Progress": ("CCE5FF", "004085"),
}

def _fill(hex_: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_)

def _font(bold=False, color="000000", size=9) -> Font:
    return Font(name="Arial", bold=bold, color=color, size=size)

def _aln(h="left", v="center", wrap=True) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ── Column map (1-based) matching the JIRA import sheet ──────────────────────
COL = {
    "tc_id":          1,   # A
    "summary":        2,   # B
    "module":         3,   # C
    "priority":       4,   # D
    "labels":         5,   # E
    "precondition":   6,   # F
    "step_action":    7,   # G
    "step_data":      8,   # H
    "step_expected":  9,   # I
    "exec_status":   10,   # J
    "actual_result": 11,   # K
    "exec_date":     12,   # L
    "exec_time":     13,   # M
    "executed_by":   14,   # N
    "defect_id":     15,   # O
    "comments":      16,   # P
}


class ExcelManager:
    """Manages reading test data and writing results to the Excel workbook."""

    SHEET_IMPORT = "🧪 Test Cases (Import)"
    SHEET_DATA   = "TestData"           # optional dedicated data sheet

    def __init__(self, result_file: Path = EXCEL_RESULT_FILE):
        self.result_file = Path(result_file)

    # ── Read ──────────────────────────────────────────────────────────────────
    def get_test_data(self, tc_id: str) -> dict:
        """
        Return a dict of test data for a given TC ID from the import sheet.
        Falls back to empty strings if the TC is not found.
        """
        if not self.result_file.exists():
            raise FileNotFoundError(f"Excel not found: {self.result_file}")

        wb = load_workbook(str(self.result_file), data_only=True)
        ws = wb[self.SHEET_IMPORT]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[COL["tc_id"] - 1] == tc_id:
                return {
                    "tc_id":         row[COL["tc_id"]        - 1] or "",
                    "summary":       row[COL["summary"]       - 1] or "",
                    "module":        row[COL["module"]        - 1] or "",
                    "priority":      row[COL["priority"]      - 1] or "",
                    "labels":        row[COL["labels"]        - 1] or "",
                    "precondition":  row[COL["precondition"]  - 1] or "",
                    "step_action":   row[COL["step_action"]   - 1] or "",
                    "step_data":     row[COL["step_data"]     - 1] or "",
                    "step_expected": row[COL["step_expected"] - 1] or "",
                }
        raise ValueError(f"TC ID '{tc_id}' not found in {self.result_file.name}")

    def get_all_tc_ids(self) -> list[str]:
        """Return all TC IDs from the import sheet (used by the full suite runner)."""
        wb = load_workbook(str(self.result_file), data_only=True)
        ws = wb[self.SHEET_IMPORT]
        return [
            row[0] for row in ws.iter_rows(min_row=2, values_only=True)
            if row[0] and str(row[0]).startswith("TC-")
        ]

    # ── Write ─────────────────────────────────────────────────────────────────
    def write_result(
        self,
        tc_id:        str,
        status:       str,
        actual:       str,
        exec_time:    str,
        tester:       str,
        defect_id:    str = "",
        comments:     str = "",
        exec_date:    Optional[str] = None,
    ) -> None:
        """Update execution columns J-P for the given TC ID."""
        if not self.result_file.exists():
            print(f"  ⚠️  Excel not found at {self.result_file} — skipping write.")
            return

        if exec_date is None:
            exec_date = datetime.now().strftime("%Y-%m-%d")

        wb = load_workbook(str(self.result_file))
        ws = wb[self.SHEET_IMPORT]
        bg_hex, fg_hex = STATUS_STYLE.get(status, ("FFFFFF", "000000"))

        for row in ws.iter_rows(min_row=2):
            if row[COL["tc_id"] - 1].value == tc_id:
                updates = {
                    COL["exec_status"]:   (status,    True,  bg_hex,   fg_hex,   "center"),
                    COL["actual_result"]: (actual,    False, "F5F5F5", "222222", "left"),
                    COL["exec_date"]:     (exec_date, False, "F5F5F5", "222222", "center"),
                    COL["exec_time"]:     (exec_time, False, "F5F5F5", "222222", "center"),
                    COL["executed_by"]:   (tester,    False, "F5F5F5", "222222", "left"),
                    COL["defect_id"]:     (defect_id, False, "F5F5F5", "222222", "center"),
                    COL["comments"]:      (comments,  False, "F5F5F5", "222222", "left"),
                }
                for col_idx, (val, bold, bg, fg, align) in updates.items():
                    cell            = row[col_idx - 1]
                    cell.value      = val
                    cell.font       = _font(bold=bold, color=fg)
                    cell.fill       = _fill(bg)
                    cell.alignment  = _aln(h=align)
                    cell.border     = THIN_BORDER
                break

        wb.save(str(self.result_file))
        print(f"  📝  Excel updated → {self.result_file.name}  [{tc_id} = {status}]")
